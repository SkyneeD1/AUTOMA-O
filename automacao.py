
# -*- coding: utf-8 -*-
"""
Automação eLaw - Cadastro / Atualização com planilha
VERSÃO: V3.3.10 (DATA NORMALIZER + modo humano + IFRAME FIX + OPEN EXCEL ON ERROR)

• Datas normalizadas antes de digitar (evita 5040/5041, 16/10/2025 aleatório, etc.)
• Modo humano para datas (digitação lenta + ENTER real)
• Modais PrimeFaces com IFRAME (Juiz e Parte Contrária) – preenche via JS dentro do iframe
• STATUS com dtype object (sem FutureWarning)
• Abre Excel automaticamente se houver linhas com erro (amarelas)
"""

import os
import re
import time
import math
import traceback
from datetime import datetime, timedelta

import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service

# =====================
# TYPE HINTS (opcionais)
# =====================
from typing import Callable, Optional


from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =====================
# CONFIGURAÇÕES
# =====================
EXCEL_PATH = "PLANILHA CADASTRO NOVA AÇÃO.xlsx"
CHROMEDRIVER_PATH = "C:/chromedriver/chromedriver.exe"  # ajuste conforme ambiente
SITE_URL = "https://vtal.elaw.com.br/"
YELLOW_HEX = "FFF200"
WAIT_SHORT = 8
WAIT_MEDIUM = 20
WAIT_LONG = 40

# =====================
# LER PLANILHA
# =====================
# Lemos sem forçar dtype para que datas em número (serial Excel) sejam detectáveis;
# a normalização cuida de todos os formatos.
df = pd.read_excel(EXCEL_PATH)
if "STATUS" not in df.columns:
    df["STATUS"] = ""
# Garante dtype texto p/ evitar FutureWarning ao atribuir strings
df["STATUS"] = df["STATUS"].astype("object")

def set_status(idx, text):
    try:
        df.at[idx, "STATUS"] = str(text)
    except Exception:
        df.loc[idx, "STATUS"] = str(text)

# Colunas (nomes conforme sua planilha)
COL_NUM_PROCESSO         = "Número do processo"
COL_RITO                 = "Localização do Processo"
COL_ESTADO               = "Estado"
COL_COMARCA              = "Comarca"
COL_FORO                 = "Foro/Tribunal"
COL_VARA                 = "Vara"
COL_CLASSIFICACAO        = "Classificação Interna"
COL_INSTANCIA            = "Instância"
COL_FASE                 = "Fase"
COL_JUIZ                 = "Juiz"
COL_CLIENTE_EMPRESA      = "Empresa e Forma de participação"
COL_CPF_PARTE_CONTR      = "CPF DA PARTE CONTRARIA"
COL_EMPREGADORA          = "Empregadora"
COL_TIPO_EMPREGADO       = "Tipo Empregado"
COL_ADV_CONTR            = "Advogado da Parte Contrária"
COL_DATA_DISTR           = "Data de Distribuição"
COL_DATA_CITACAO         = "Data de Citação"
COL_TIPO_ACAO            = "Tipo de Ação"
COL_VALOR_CAUSA          = "Valor da Causa"
COL_ADV_RESP             = "Advogado Responsável"
COL_GESTOR_JURIDICO      = "Gestor Jurídico"
COL_TIPO_DOC             = "Tipo de Documento"


rows_to_color_yellow = set()

# =====================
# NORMALIZAÇÃO DE DATAS (robusta)
# =====================
EXCEL_EPOCH = datetime(1899, 12, 30)  # Regra do Excel (considerando bug do 29/02/1900)

def as_ddmmyyyy(raw):
    """
    Converte qualquer 'raw' (string, número serial do Excel, datetime, etc.) em 'DD/MM/YYYY'.
    Retorna "" se não for possível.
    """
    if raw is None:
        return ""
    # Se vier do pandas como NaT/NaN
    try:
        if pd.isna(raw):
            return ""
    except Exception:
        pass

    # Caso já seja datetime
    if isinstance(raw, (datetime, pd.Timestamp)):
        return raw.strftime("%d/%m/%Y")

    # Caso seja número -> tentar como serial do Excel
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        if math.isfinite(raw):
            # número pequeno provavelmente não é serial válido; ainda assim tentamos coerção segura
            try:
                dt = EXCEL_EPOCH + timedelta(days=float(raw))
                # sanity check: ano entre 1900 e 2100
                if 1900 <= dt.year <= 2100:
                    return dt.strftime("%d/%m/%Y")
            except Exception:
                pass

    # Trata como string
    s = str(raw).strip()
    if not s:
        return ""

    # Tenta parsing com dayfirst e com monthfirst
    for dayfirst in (True, False):
        try:
            dt = pd.to_datetime(s, dayfirst=dayfirst, errors="raise")
            # sanity check:
            if 1900 <= dt.year <= 2100:
                return dt.strftime("%d/%m/%Y")
        except Exception:
            pass

    # Tenta formatos explícitos
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            if 1900 <= dt.year <= 2100:
                return dt.strftime("%d/%m/%Y")
        except Exception:
            continue

    # Última tentativa: apenas números tipo DDMMYYYY ou YYYYMMDD
    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        # Tentativa DDMMYYYY
        try:
            dt = datetime.strptime(digits, "%d%m%Y")
            return dt.strftime("%d/%m/%Y")
        except Exception:
            pass
        # Tentativa YYYYMMDD
        try:
            dt = datetime.strptime(digits, "%Y%m%d")
            return dt.strftime("%d/%m/%Y")
        except Exception:
            pass

    return ""

# =====================
# SELENIUM SETUP
# =====================
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
# options.add_argument("--headless=new")  # se quiser headless
service = Service(CHROMEDRIVER_PATH)
driver = webdriver.Chrome(service=service, options=options)
wait = WebDriverWait(driver, WAIT_LONG)

# =====================
# HELPERS
# =====================
def safe_text(val):
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    return str(val).strip()

def to_amount_str(val):
    if val is None or (isinstance(val, float) and math.isnan(val)) or (isinstance(val, str) and not val.strip()):
        return ""
    try:
        # normaliza "1.234,56" -> "1234.56"
        return str(float(str(val).replace(".", "").replace(",", ".")))
    except Exception:
        return str(val).replace(",", ".")


def tentar_selecionar_primeiro_item_autocomplete(painel_id: str):
    """Tenta clicar diretamente no primeiro item do autocomplete informado.

    Retorna o label do item selecionado quando bem-sucedido, caso contrário False.
    """
    if not painel_id:
        return False

    try:
        painel_wait = WebDriverWait(driver, WAIT_SHORT)
        painel_wait.until(EC.visibility_of_element_located((By.ID, painel_id)))
        xpath_primeiro_item = (
            f"//*[@id={_xpath_literal(painel_id)}]//li[contains(@class,'ui-autocomplete-item')]"
        )
        primeiro_item = painel_wait.until(
            EC.element_to_be_clickable((By.XPATH, xpath_primeiro_item))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'nearest'});", primeiro_item)
        label = (primeiro_item.get_attribute("data-item-label") or primeiro_item.text or "").strip()
        try:
            primeiro_item.click()
        except Exception:
            driver.execute_script("arguments[0].click();", primeiro_item)

        try:
            painel_wait.until(EC.invisibility_of_element_located((By.ID, painel_id)))
        except Exception:
            pass

        time.sleep(0.2)
        return label or True
    except Exception as e:
        print(f"ℹ️ Não foi possível clicar no primeiro item do autocomplete {painel_id}: {e}")
        return False


def wait_element_by_id_suffix(
    suffix: str,
    tag: str = "*",
    timeout: int = WAIT_LONG,
    condition: Optional[Callable] = None,
):
    """Localiza um elemento usando o final do seu ID (suffix).

    Útil para componentes PrimeFaces com IDs dinâmicos que mudam entre telas
    (ex.: j_id_4c_* x j_id_4g_*). Permite informar o *tag* para restringir a busca
    e uma *condition* (ex.: EC.element_to_be_clickable) quando necessário.
    """

    selector = f"{tag}[id$='{suffix}']"
    locator = (By.CSS_SELECTOR, selector)
    expected = condition(locator) if condition else EC.presence_of_element_located(locator)
    return WebDriverWait(driver, timeout).until(expected)




def attempt_twice(action_desc, func, *args, **kwargs):
    for tent in range(1, 2 + 1):
        try:
            r = func(*args, **kwargs)
            print(f"✅ {action_desc} (tentativa {tent})")
            return True if r is None else r
        except Exception as e:
            print(f"⚠️ Falha em '{action_desc}' (tentativa {tent}): {e}")
            if tent == 1:
                time.sleep(1.2)
    return False


def _xpath_literal(texto: str) -> str:
    """Escapa corretamente strings para uso em XPaths (lida com aspas simples/duplas)."""
    if "'" not in texto:
        return f"'{texto}'"
    if '"' not in texto:
        return f'"{texto}"'
    partes = texto.split("'")
    pedacos = []
    for idx, parte in enumerate(partes):
        if parte:
            pedacos.append(f"'{parte}'")
        if idx != len(partes) - 1:
            pedacos.append("\"'\"")
    return "concat(" + ",".join(pedacos) + ")"

def clicar_id(elem_id):
    print(f"➡️ Clicar ID: {elem_id}")
    elem = wait.until(EC.element_to_be_clickable((By.ID, elem_id)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
    try:
        elem.click()
    except Exception:
        driver.execute_script("arguments[0].click();", elem)
    time.sleep(0.5)

def preencher_input(input_id, valor, clear_first=True, press_enter=False):
    """
    Preenchimento padrão (não datas): limpa com CTRL+A+Backspace e cola EXATO.
    """
    if valor == "" and valor != 0:
        return
    elem = wait.until(EC.presence_of_element_located((By.ID, input_id)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
    elem.click()
    time.sleep(0.2)
    elem.send_keys(Keys.CONTROL, "a")
    elem.send_keys(Keys.BACKSPACE)
    time.sleep(0.2)
    elem.send_keys(str(valor))
    if press_enter:
        time.sleep(0.2)
        elem.send_keys(Keys.ENTER)
    time.sleep(0.4)

# ✅ MODO HUMANO PARA DATAS — digita devagar e confirma com ENTER de teclado
def digitar_data_humano(input_id, data_valor):
    try:
        if not data_valor:
            return True
        campo = wait.until(EC.element_to_be_clickable((By.ID, input_id)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
        campo.click()
        time.sleep(0.25)
        campo.send_keys(Keys.CONTROL, "a")
        campo.send_keys(Keys.BACKSPACE)
        time.sleep(0.15)
        for ch in data_valor:
            campo.send_keys(ch)
            time.sleep(0.06)  # ritmo humano
        time.sleep(0.1)
        campo.send_keys(Keys.ENTER)  # confirmar
        print(f"✅ Data '{data_valor}' digitada (modo humano) em {input_id}")
        time.sleep(0.35)
        return True
    except Exception as e:
        print(f"❌ Erro ao digitar data manual em {input_id}: {e}")
        return False

def existe_xpath(xpath):
    try:
        driver.find_element(By.XPATH, xpath)
        return True
    except:
        return False

def anexar_arquivo_por_input(file_path):
    upload_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))
    upload_input.send_keys(os.path.abspath(file_path))
    time.sleep(0.8)

def marcar_erro(idx, etapa, err):
    msg = f"ERRO {etapa}: {err}"
    print(f"❌ {msg}")
    set_status(idx, f"⚠️ {msg}")
    rows_to_color_yellow.add(idx)


def esperar_texto_em_tabela_outras_partes(texto: str, timeout=WAIT_MEDIUM) -> bool:
    if not texto:
        return False
    literal = _xpath_literal(texto.strip())
    xpath = (
        "//table[contains(@id,'outrasParte') and contains(@class,'ui-datatable')]"
        f"//span[contains(normalize-space(.), {literal})]"
    )
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        return True
    except Exception as e:
        print(f"⚠️ Não encontrei '{texto}' na lista de Outras Partes: {e}")
        return False


def preencher_autocomplete_por_rotulo(
    rotulo: str,
    valor: str,
    tempo_dropdown: float = 0.9,
) -> bool:
    if not valor:
        return True
    literal = _xpath_literal(rotulo)
    input_xpath = (
        f"//label[contains(normalize-space(.), {literal})]"
        "//following::input[contains(@id,'autocomplete')][1]"
    )

    def _preencher():
        campo = wait.until(EC.presence_of_element_located((By.XPATH, input_xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
        campo.clear()
        time.sleep(0.15)
        campo.send_keys(valor)
        time.sleep(tempo_dropdown)
        campo_id = campo.get_attribute("id") or ""
        painel_id = ""
        if campo_id.endswith("_input"):
            painel_id = f"{campo_id[:-len('_input')]}_panel"
        if painel_id:
            selecionado = tentar_selecionar_primeiro_item_autocomplete(painel_id)
            if selecionado:
                esperado = str(selecionado).strip()
                if esperado:
                    try:
                        WebDriverWait(driver, WAIT_SHORT).until(
                            lambda d: esperado.lower()
                            in (campo.get_attribute("value") or "").lower()
                        )
                    except Exception:
                        pass
                try:
                    campo.send_keys(Keys.ENTER)
                except Exception:
                    pass
                time.sleep(0.4)
                return
        campo.send_keys(Keys.DOWN)
        time.sleep(0.25)
        try:
            campo.send_keys(Keys.ENTER)
        except Exception:
            pass
        time.sleep(0.4)

    if attempt_twice(
        f"Preencher '{rotulo}' com {valor}",
        _preencher,
    ):
        return True
    return False


def preencher_autocomplete_por_id(
    input_id: str,
    valor: str,
    tempo_dropdown: float = 0.9,
) -> bool:
    if not valor:
        return True
    painel_id = ""
    if input_id.endswith("_input"):
        painel_id = f"{input_id[:-len('_input')]}_panel"

    def _preencher():
        campo = wait.until(EC.presence_of_element_located((By.ID, input_id)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo)
        campo.clear()
        time.sleep(0.15)
        campo.send_keys(valor)
        time.sleep(tempo_dropdown)
        if painel_id:
            selecionado = tentar_selecionar_primeiro_item_autocomplete(painel_id)
            if selecionado:
                esperado = str(selecionado).strip()
                if esperado:
                    try:
                        WebDriverWait(driver, WAIT_SHORT).until(
                            lambda d: esperado.lower()
                            in (campo.get_attribute("value") or "").lower()
                        )
                    except Exception:
                        pass
                try:
                    campo.send_keys(Keys.ENTER)
                except Exception:
                    pass
                time.sleep(0.4)
                return
        campo.send_keys(Keys.DOWN)
        time.sleep(0.25)
        try:
            campo.send_keys(Keys.ENTER)
        except Exception:
            pass
        time.sleep(0.4)

    if attempt_twice(
        f"Preencher autocomplete {input_id} com {valor}",
        _preencher,
    ):
        return True
    return False

def colorir_linhas_amarelo_no_excel(excel_path, linhas_idx, header_rows=1):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        fill = PatternFill(start_color=YELLOW_HEX, end_color=YELLOW_HEX, fill_type="solid")
        for idx in linhas_idx:
            excel_row = idx + 1 + header_rows
            for col in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=col).fill = fill
        wb.save(excel_path)
        print(f"🎨 Linhas coloridas de amarelo: {[i+1 for i in linhas_idx]}")
    except Exception as e:
        print(f"⚠️ Falha ao colorir linhas no Excel: {e}")

# ================
# PRIMEFACES SELECT
# ================
_SIGLA_ESTADO_RE = re.compile(r"^[A-Z]{2}$")

def _ajusta_valor_para_estado(label_id: str, valor: str) -> str:
    if not valor:
        return valor
    id_lower = label_id.lower()
    pode_ser_estado = ("comboestadovara" in id_lower) or ("estado" in id_lower)
    if pode_ser_estado and _SIGLA_ESTADO_RE.match(valor.strip().upper()):
        return valor.strip().upper() + " -"
    return valor

def selecionar_primefaces(label_id, valor, timeout=WAIT_LONG):
    valor = _ajusta_valor_para_estado(label_id, (valor or "").strip())
    label = wait.until(EC.element_to_be_clickable((By.ID, label_id)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", label)
    driver.execute_script("arguments[0].click();", label)
    time.sleep(0.25)
    panel = WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((
            By.XPATH,
            "//div[contains(@class,'ui-selectonemenu-panel') and contains(@style,'display: block')]"
        ))
    )
    try:
        filtro = panel.find_element(By.XPATH, ".//input[contains(@id,'_filter')]")
        filtro.clear()
        if valor:
            filtro.send_keys(valor)
        time.sleep(0.6)
        filtro.send_keys(Keys.ENTER)
        time.sleep(0.35)
        return True
    except Exception:
        js = ("var p=document.querySelector(\"div.ui-selectonemenu-panel[style*='display: block'] li:not(.ui-state-disabled)\");"
              "if(p){p.click(); return true;} return false;")
        ok = driver.execute_script(js)
        if ok:
            time.sleep(0.25)
            return True
        raise Exception(f"Não foi possível selecionar no dropdown {label_id}")

# =====================
# MODAIS COM IFRAME (Juiz + Parte Contrária)
# =====================
def _get_visible_dialogs():
    dialogs = driver.find_elements(By.CSS_SELECTOR, "div.ui-dialog.ui-overlay-visible, div.ui-dialog[style*='display: block']")
    return [d for d in dialogs if d.is_displayed()]

def _find_dialog_iframe(dialog):
    try:
        return dialog.find_element(By.CSS_SELECTOR, "iframe")
    except Exception:
        return None

def _switch_into_dialog_iframe_by_hint(id_hint_contains: str, timeout=WAIT_LONG):
    """
    Tenta achar o dialog visível cujo id contém 'id_hint_contains' e entrar no seu iframe.
    Se não achar por hint, pega qualquer dialog visível com iframe.
    Retorna o WebElement do dialog usado (para aguardar o fechamento depois).
    """
    t0 = time.time()
    while time.time() - t0 < timeout:
        dialogs = _get_visible_dialogs()
        # 1) tenta por hint
        for d in dialogs:
            try:
                dlg_id = d.get_attribute("id") or ""
                if id_hint_contains and id_hint_contains in dlg_id:
                    ifr = _find_dialog_iframe(d)
                    if ifr:
                        print(f"🔎 Dialog alvo encontrado (id='{dlg_id}'), entrando no iframe...")
                        driver.switch_to.frame(ifr)
                        return d
            except Exception:
                pass
        # 2) pega qualquer dialog com iframe
        for d in dialogs:
            ifr = _find_dialog_iframe(d)
            if ifr:
                dlg_id = d.get_attribute("id") or ""
                print(f"🔎 Dialog visível com iframe encontrado (id='{dlg_id}'), entrando no iframe...")
                driver.switch_to.frame(ifr)
                return d
        time.sleep(0.2)
    raise Exception("Timeout ao localizar iframe dentro de um dialog visível.")

def _leave_iframe():
    try:
        driver.switch_to.default_content()
    except Exception:
        pass

def _wait_dialog_invisible(dialog, timeout=WAIT_LONG):
    try:
        WebDriverWait(driver, timeout).until(EC.invisibility_of_element_located((By.ID, dialog.get_attribute("id"))))
        return True
    except Exception:
        return False

def criar_juiz_modal_js(juiz_nome: str):
    """
    Abre modal de Juiz, entra no iframe, preenche j_id_w, clica salvar (btnSalvarjuiz) e aguarda fechar.
    """
    print("➡️ Abrindo modal Juiz (Novo)...")
    clicar_id("j_id_4c_1:juizBtnNovo")

    print("⏳ Aguardando dialog + iframe do Juiz...")
    dialog = _switch_into_dialog_iframe_by_hint("juizBtnNovo_dlg", timeout=WAIT_LONG)

    try:
        print("✍️ Preenchendo nome do Juiz (input#j_id_w)...")
        input_elem = WebDriverWait(driver, WAIT_LONG).until(
            EC.presence_of_element_located((By.ID, "j_id_w"))
        )
        driver.execute_script("arguments[0].focus();", input_elem)
        driver.execute_script("arguments[0].value = arguments[1];", input_elem, juiz_nome)
        driver.execute_script("arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", input_elem)
        driver.execute_script("arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", input_elem)
        time.sleep(0.3)

        print("💾 Clicando salvar do Juiz...")
        try:
            salvar_btn = driver.find_element(By.ID, "btnSalvarjuiz")
        except Exception:
            salvar_btn = driver.find_element(By.CSS_SELECTOR, "button[id*='Salvar']")
        driver.execute_script("arguments[0].click();", salvar_btn)
        time.sleep(0.8)
    except Exception as e:
        raise Exception(f"Erro ao preencher/salvar Juiz dentro do iframe: {e}")
    finally:
        print("↩️ Retornando para o contexto principal...")
        _leave_iframe()

    if _wait_dialog_invisible(dialog, timeout=WAIT_LONG):
        print("✅ Modal Juiz fechado.")
    else:
        print("⚠️ Modal Juiz ainda visível, prosseguindo (pode ser renderização tardia).")

def incluir_parte_contraria_modal_js(cpf_cnpj: str):
    """
    Abre modal de Parte Contrária, entra no iframe, preenche CPF/CNPJ (input#j_id_1e),
    clica Continuar (button#j_id_1i), depois Salvar (button#parteContrariaButtom), aguarda fechar.
    """
    print("➡️ Abrindo modal Parte Contrária (Novo)...")
    clicar_id("j_id_4c_1:j_id_4c_5_2_2_b_9_8_1:parteContrariaMainGridBtnNovo")

    print("⏳ Aguardando dialog + iframe da Parte Contrária...")
    dialog = _switch_into_dialog_iframe_by_hint("parteContrariaMainGridBtnNovo_dlg", timeout=WAIT_LONG)

    try:
        print("✍️ Preenchendo CPF/CNPJ (input#j_id_1e)...")
        input_elem = WebDriverWait(driver, WAIT_LONG).until(
            EC.presence_of_element_located((By.ID, "j_id_1e"))
        )
        driver.execute_script("arguments[0].focus();", input_elem)
        driver.execute_script("arguments[0].value = arguments[1];", input_elem, cpf_cnpj)
        driver.execute_script("arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", input_elem)
        driver.execute_script("arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", input_elem)
        time.sleep(0.3)

        print("➡️ Clicando 'Continuar' (button#j_id_1i)...")
        try:
            cont_btn = driver.find_element(By.ID, "j_id_1i")
        except Exception:
            # fallback
            try:
                cont_btn = driver.find_element(By.CSS_SELECTOR, "button[id*='1i'], button[id*='Continuar']")
            except Exception as e2:
                raise Exception(f"Botão Continuar não localizado: {e2}")
        driver.execute_script("arguments[0].click();", cont_btn)
        time.sleep(0.8)

        print("💾 Aguardando e clicando 'Salvar' da Parte Contrária (button#parteContrariaButtom)...")
        try:
            save_btn = WebDriverWait(driver, WAIT_LONG).until(
                EC.visibility_of_element_located((By.ID, "parteContrariaButtom"))
            )
        except Exception:
            try:
                save_btn = driver.find_element(By.CSS_SELECTOR, "button[id*='parteContraria'], button[id*='Salvar']")
            except Exception as e3:
                raise Exception(f"Botão Salvar da Parte Contrária não apareceu: {e3}")
        driver.execute_script("arguments[0].click();", save_btn)
        time.sleep(0.8)
    except Exception as e:
        raise Exception(f"Erro ao incluir Parte Contrária dentro do iframe: {e}")
    finally:
        print("↩️ Retornando para o contexto principal...")
        _leave_iframe()

    if _wait_dialog_invisible(dialog, timeout=WAIT_LONG):
        print("✅ Modal Parte Contrária fechado.")
    else:
        print("⚠️ Modal Parte Contrária ainda visível, prosseguindo (pode ser renderização tardia).")

# =====================
# FLUXO PRINCIPAL
# =====================
try:
    driver.get(SITE_URL)
    print("👀 Aguardando login... (até 180s)")
    try:
        WebDriverWait(driver, 180).until(EC.url_contains("/homePage.elaw"))
        print("✅ Login detectado, iniciando automação...")
    except:
        print("⚠️ Login não detectado automaticamente. Faça login e pressione ENTER aqui.")
        input("👉 Pressione ENTER após logar...")

    for idx, row in df.iterrows():
        processo = safe_text(row.get(COL_NUM_PROCESSO, ""))
        if not processo:
            continue

        print("\n" + "="*86)
        print(f"🔎 Linha {idx+1} | Processo: {processo}")
        set_status(idx, "EM ANDAMENTO...")

        # extrair campos
        rito            = safe_text(row.get(COL_RITO, ""))
        estado_vara     = safe_text(row.get(COL_ESTADO, ""))
        comarca_vara    = safe_text(row.get(COL_COMARCA, ""))
        foro_tribunal   = safe_text(row.get(COL_FORO, ""))
        vara_especifica = safe_text(row.get(COL_VARA, ""))
        classificacao   = safe_text(row.get(COL_CLASSIFICACAO, ""))
        instancia       = safe_text(row.get(COL_INSTANCIA, ""))
        fase_processo   = safe_text(row.get(COL_FASE, ""))
        juiz_nome       = safe_text(row.get(COL_JUIZ, ""))
        cliente_empresa = safe_text(row.get(COL_CLIENTE_EMPRESA, ""))
        cpf_cnpj_contr  = safe_text(row.get(COL_CPF_PARTE_CONTR, ""))
        empresa_nivel1  = safe_text(row.get(COL_EMPREGADORA, ""))
        tipo_parte      = safe_text(row.get(COL_TIPO_EMPREGADO, ""))
        advogado_contr  = safe_text(row.get(COL_ADV_CONTR, ""))
        tipo_processo   = safe_text(row.get(COL_TIPO_ACAO, ""))
        valor_causa     = to_amount_str(row.get(COL_VALOR_CAUSA, ""))
        adv_resp        = safe_text(row.get(COL_ADV_RESP, ""))
        gestor_juridico = safe_text(row.get(COL_GESTOR_JURIDICO, ""))

        # DATAS normalizadas (robustas)
        data_distrib    = as_ddmmyyyy(row.get(COL_DATA_DISTR, ""))
        data_receb      = as_ddmmyyyy(row.get(COL_DATA_CITACAO, ""))

        tipo_doc_val    = safe_text(row.get(COL_TIPO_DOC, "")) or "Petição Inicial"

        pdf_filename = f"ATOrd_{processo}.pdf"
        pdf_path = os.path.join(os.getcwd(), pdf_filename)

        try:
            # abrir processo via autocomplete global
            def _abrir_processo():
                search_input = WebDriverWait(driver, WAIT_LONG).until(
                    EC.presence_of_element_located((By.ID, "j_id_2g:globaSearchAutocomplete_input"))
                )
                search_input.clear()
                time.sleep(0.25)
                search_input.send_keys(processo)
                WebDriverWait(driver, WAIT_MEDIUM).until(
                    EC.visibility_of_element_located((By.XPATH, f"//span[contains(text(),'{processo}')]"))
                )
                time.sleep(0.4)
                search_input.send_keys(Keys.DOWN)
                time.sleep(0.25)
                search_input.send_keys(Keys.ENTER)
                time.sleep(0.8)
            if not attempt_twice("Abrir processo pelo autocomplete", _abrir_processo):
                raise Exception("Não foi possível abrir o processo.")

            # entrar no modo editar
            if not attempt_twice("Entrar no modo Editar", clicar_id, "btnEditar"):
                raise Exception("Botão Editar indisponível.")

            # DROPDOWNS
            if rito:
                attempt_twice("Selecionar Rito", selecionar_primefaces,
                              "j_id_4c_1:j_id_4c_5_2_2_1_9_t_1:comboRito_label", rito)
            if estado_vara:
                attempt_twice("Selecionar Estado", selecionar_primefaces,
                              "j_id_4c_1:j_id_4c_5_2_2_1_9_t_1:comboEstadoVara_label", estado_vara)
            if comarca_vara:
                attempt_twice("Selecionar Comarca", selecionar_primefaces,
                              "j_id_4c_1:j_id_4c_5_2_2_1_9_t_1:comboComarcaVara_label", comarca_vara)
            if foro_tribunal:
                attempt_twice("Selecionar Foro/Tribunal", selecionar_primefaces,
                              "j_id_4c_1:j_id_4c_5_2_2_1_9_t_1:comboForoTribunal_label", foro_tribunal)
            if vara_especifica:
                attempt_twice("Selecionar Vara", selecionar_primefaces,
                              "j_id_4c_1:j_id_4c_5_2_2_1_9_t_1:comboVara_label", vara_especifica)
            if classificacao:
                attempt_twice("Selecionar Classificação", selecionar_primefaces,
                              "j_id_4c_1:j_id_4c_5_2_2_2_9_15_1:processoClassificacaoCombo_label", classificacao)
            if instancia:
                attempt_twice("Selecionar Instância", selecionar_primefaces,
                              "j_id_4c_1:j_id_4c_5_2_2_3_9_19_1_label", instancia)
            if fase_processo:
                attempt_twice("Selecionar Fase", selecionar_primefaces,
                              "j_id_4c_1:processoFaseCombo_label", fase_processo)
            if cliente_empresa:
                attempt_twice("Selecionar Empresa (Cliente)", selecionar_primefaces,
                              "j_id_4c_1:comboClientProcessoParte_label", cliente_empresa)

            # Papel = Réu
            attempt_twice("Selecionar Papel = Réu", selecionar_primefaces,
                          "j_id_4c_1:j_id_4c_5_2_2_9_9_2_6_label", "Réu")

            # Tipo de documento
            if tipo_doc_val:
                attempt_twice("Selecionar Tipo de Documento", selecionar_primefaces,
                              "j_id_4c_1:j_id_4c_5_2_2_r_9_24_1:eFileTipoCombo_label", tipo_doc_val)

            # Parte do documento = Autor
            attempt_twice("Selecionar Parte do Documento = Autor", selecionar_primefaces,
                          "j_id_4c_1:j_id_4c_5_2_2_b_9_8_1:j_id_4c_5_2_2_b_9_8_5_2_n_label", "Autor")

            # JUIZ modal (iframe)
            if juiz_nome:
                if not attempt_twice("Criar Juiz (Modal c/ iframe)", criar_juiz_modal_js, juiz_nome):
                    raise Exception("Juiz não pôde ser criado via modal.")

            # PARTE CONTRÁRIA modal (iframe)
            if cpf_cnpj_contr:
                if not attempt_twice("Incluir Parte Contrária (Modal c/ iframe)", incluir_parte_contraria_modal_js, cpf_cnpj_contr):
                    raise Exception("Falha ao incluir parte contrária via modal.")

            # Advogado parte contrária (autocomplete)
            if advogado_contr:
                def _adv_contra():
                    inp = wait.until(EC.presence_of_element_located((By.ID, "j_id_4c_1:j_id_4c_5_2_2_f_9_2v_1:autocompleteAdvogadoParteContrariaNome_input")))
                    inp.clear()
                    time.sleep(0.15)
                    inp.send_keys(advogado_contr)
                    time.sleep(0.9)
                    inp.send_keys(Keys.DOWN)
                    time.sleep(0.2)
                    inp.send_keys(Keys.ENTER)
                    time.sleep(0.4)
                attempt_twice("Selecionar Advogado da Parte Contrária", _adv_contra)

            # ✅ DATAS com normalização + digitação humana
            if data_distrib:
                attempt_twice("DIGITAR Data Distribuição (humano)", digitar_data_humano,
                              "j_id_4c_1:dataDistribuicao_input", data_distrib)

            if data_receb:
                attempt_twice("DIGITAR Data Citação (humano)", digitar_data_humano,
                              "j_id_4c_1:dataRecebimento_input", data_receb)

            # Tipo de ação
            if tipo_processo:
                attempt_twice("Selecionar Tipo de Ação", selecionar_primefaces,
                              "j_id_4c_1:comboProcessoTipo_label", tipo_processo)

            # Valor da causa
            if valor_causa:
                attempt_twice("Preencher Valor da Causa", preencher_input,
                              "j_id_4c_1:amountCase_input", valor_causa)

            # Advogado responsável (autocomplete + selectOneMenu)
            if adv_resp:
                adv_resp_input_id = "j_id_4c_1:autoCompleteLawyer_input"
                if not preencher_autocomplete_por_id(adv_resp_input_id, adv_resp):
                    print("⚠️ Autocomplete de Advogado Responsável não retornou resultados válidos.")
                else:
                    attempt_twice(
                        "Selecionar Advogado Responsável",
                        selecionar_primefaces,
                        "j_id_4c_1:comboAdvogadoResponsavelProcesso_label",
                        adv_resp,
                    )

            # Gestor Jurídico (autocomplete específico)
            if gestor_juridico:
                gestor_input_id = (
                    "j_id_4c_1:j_id_4c_5_2_2_l_9_45_2:j_id_4c_5_2_2_l_9_45_3_1_2_2_1_1:"
                    "j_id_4c_5_2_2_l_9_45_3_1_2_2_1_2g_input"
                )
                if not preencher_autocomplete_por_id(
                    gestor_input_id,
                    gestor_juridico,
                ):
                    print("⚠️ Campo 'Gestor Jurídico' não foi atualizado automaticamente.")

            # =========================
            # ✅ INCLUSÃO DE OUTRAS RECLAMADAS (1ª → 7ª RECLAMADA)
            # =========================
            colunas_reclamadas = [
                "1ª Reclamada", "2ª Reclamada", "3ª Reclamada",
                "4ª Reclamada", "5ª Reclamada", "6ª Reclamada", "7ª Reclamada"
            ]

            reclamadas_nomes = [safe_text(row.get(col, "")) for col in colunas_reclamadas]

            for parte_nome in reclamadas_nomes:
                if not parte_nome or parte_nome.strip() == "":
                    continue  # Se célula vazia, apenas passa pra próxima

                print(f"➕ Adicionando reclamada adicional: {parte_nome}")

                try:
                    # 1. AUTOCOMPLETE - DIGITAR NOME E SELECIONAR NO DROPDOWN
                    def _preencher_autocomplete_parte():
                        inp = wait_element_by_id_suffix(
                            ":autocompleteOutraParte_input",
                            tag="input",
                            condition=EC.element_to_be_clickable,
                        )
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
                        inp.click()
                        time.sleep(0.15)
                        inp.send_keys(Keys.CONTROL, "a")
                        inp.send_keys(Keys.BACKSPACE)
                        time.sleep(0.1)
                        inp.send_keys(parte_nome)

                        painel_id = inp.get_attribute("aria-controls") or ""
                        if not painel_id:
                            raise Exception("Autocomplete sem aria-controls (painel não identificado).")

                        panel = WebDriverWait(driver, WAIT_MEDIUM).until(
                            EC.visibility_of_element_located((By.ID, painel_id))
                        )

                        primeiro_item = WebDriverWait(driver, WAIT_MEDIUM).until(
                            EC.element_to_be_clickable((
                                By.CSS_SELECTOR,
                                f"#{painel_id} li.ui-autocomplete-item:not(.ui-state-disabled)",
                            ))
                        )

                        label_item = (primeiro_item.get_attribute("data-item-label") or primeiro_item.text or "").strip()
                        if not label_item:
                            raise Exception("Nenhum item disponível no autocomplete para a parte informada.")

                        driver.execute_script("arguments[0].scrollIntoView({block:'nearest'});", primeiro_item)
                        time.sleep(0.15)

                        # Segue o fluxo humano: seta para baixo + ENTER
                        inp.send_keys(Keys.DOWN)
                        time.sleep(0.25)
                        inp.send_keys(Keys.ENTER)

                        try:
                            WebDriverWait(driver, WAIT_SHORT).until(
                                EC.invisibility_of_element_located((By.ID, painel_id))
                            )
                        except Exception:
                            pass

                        selecionado = (inp.get_attribute("value") or "").strip()
                        if not selecionado:
                            raise Exception("Autocomplete não preencheu o campo da parte.")

                        label_lower = label_item.lower()
                        selecionado_lower = selecionado.lower()
                        parte_lower = parte_nome.lower()
                        if (
                            parte_lower not in label_lower
                            and parte_lower not in selecionado_lower
                            and selecionado_lower not in label_lower
                        ):
                            print(
                                f"ℹ️ Alerta: item selecionado '{selecionado}' difere da busca '{parte_nome}'."
                            )

                    if not attempt_twice(
                        f"Selecionar parte {parte_nome} via autocomplete",
                        _preencher_autocomplete_parte,
                    ):
                        raise Exception("Autocomplete não retornou resultados válidos.")

                    # 2. Selecionar papel = RÉU
                    def _selecionar_papel_reu():
                        label_elem = wait_element_by_id_suffix(
                            ":processoParteSelect_label",
                            tag="span",
                            condition=EC.element_to_be_clickable,
                        )
                        selecionar_primefaces(label_elem.get_attribute("id"), "Réu")

                    if not attempt_twice(
                        f"Selecionar papel = Réu para {parte_nome}",
                        _selecionar_papel_reu,
                    ):
                        raise Exception("Não foi possível definir papel = Réu.")

                    # 3. Clicar em ADICIONAR
                    def _clicar_botao_adicionar():
                        botao = wait_element_by_id_suffix(
                            ":outrasParteAddButtom",
                            tag="button",
                            condition=EC.element_to_be_clickable,
                        )
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", botao)
                        try:
                            botao.click()
                        except Exception:
                            driver.execute_script("arguments[0].click();", botao)
                        time.sleep(0.4)

                    if not attempt_twice(
                        f"Confirmar inclusão de {parte_nome}",
                        _clicar_botao_adicionar,
                    ):
                        raise Exception("Botão de adicionar não respondeu.")

                    if not esperar_texto_em_tabela_outras_partes(parte_nome):
                        raise Exception("Nome não apareceu na lista após adicionar.")

                    print(f"✅ Reclamada '{parte_nome}' adicionada com sucesso!")

                except Exception as e_parte:
                    print(f"⚠️ Falha ao adicionar {parte_nome}: {e_parte}")
                    continue  # Não para o fluxo, apenas segue para a próxima

            # UPLOAD PDF
            if not os.path.exists(pdf_path):
                print(f"⚠️ PDF não encontrado: {pdf_path}. Tentando anexar mesmo assim (verifique).")
            attempt_twice("Anexar PDF ATOrd_<processo>", anexar_arquivo_por_input, pdf_path)

            # SALVAR
            if not attempt_twice("Salvar alterações", clicar_id, "btnSalvarOpen"):
                raise Exception("Falha ao salvar (btnSalvarOpen).")

            set_status(idx, "OK")
            print(f"✅ Finalizado com sucesso: {processo}")

        except Exception as e_row:
            marcar_erro(idx, "PROCESSAMENTO LINHA", e_row)
            traceback.print_exc()

        time.sleep(0.6)

    # salvar status no excel
    df.to_excel(EXCEL_PATH, index=False)
    print("📁 Excel atualizado com STATUS.")

    # pintar linhas com erro + abrir planilha automaticamente se houver erro
    if rows_to_color_yellow:
        colorir_linhas_amarelo_no_excel(EXCEL_PATH, rows_to_color_yellow, header_rows=1)
        try:
            print("⚠️ Erros encontrados. Abrindo planilha para revisão...")
            os.startfile(EXCEL_PATH)  # Windows
        except Exception as e:
            print(f"ℹ️ Não foi possível abrir a planilha automaticamente: {e}")

except Exception as e_main:
    print(f"❌ ERRO GERAL: {e_main}")
    traceback.print_exc()
finally:
    try:
        driver.quit()
    except:
        pass
    print("🧹 Navegador encerrado.")
